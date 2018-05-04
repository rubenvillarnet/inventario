#! python3
#leerStock.py - Lee los datos del archivo de stocks
import openpyxl, sqlite3, os

def obtener_nombres(hoja):
    primera_fila = list(hoja.rows)[0]


    for campo in primera_fila:
        if campo.value == 'REF VT':
            col_ref = campo.column
        if campo.value == 'STOCK VT':
            col_stock = campo.column

    return col_ref, col_stock


def conectar_bd(db_file):
    con = sqlite3.connect(db_file)
    return con


def vaciar_tabla_inventario(con):
    sql = 'DELETE FROM inventario_articulo'
    cur = con.cursor()
    cur.execute(sql)
    con.commit()

def insertar_articulo(ref,cnt):
    cur = con.cursor()
    cur.execute(""" INSERT INTO inventario_articulo(referencia, cantidad)
        VALUES(?,?)""", (ref, cnt))
    con.commit()


#Abrir libro
#libro = openpyxl.load_workbook("C:\TEMP\stockLite.xlsx",  data_only=True)
libro = openpyxl.load_workbook(os.environ['STOCK_DIR']+"stockLite.xlsx",  data_only=True)
#libro = openpyxl.load_workbook("C:\TEMP\STOCK 2018 (III).xlsx",  data_only=True)

#Abrir hoja
hoja = libro["Hoja1"]
#crear diccionario para guardar los datos
col_ref, col_stock = obtener_nombres(hoja)
inventario = {}
con = conectar_bd(os.environ['BBDD'])
vaciar_tabla_inventario(con)

for fila in range(2, hoja.max_row + 1):
    ref= hoja[col_ref + str(fila)].value
    cnt= hoja[col_stock + str(fila)].value
    if ref != None and str(ref)[0:3] != "Tot" and str(ref)[0:3] != "TOT" and cnt != None:
        inventario.setdefault(ref, 0)
    #TODO hacer que sólo coja los valores con número
        inventario[ref] += int(cnt)

for i in inventario:
    insertar_articulo(i,inventario[i])
